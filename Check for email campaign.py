import os 
from langchain_community.tools.tavily_search import TavilySearchResults
from dotenv import load_dotenv
load_dotenv()
from langchain_groq import ChatGroq
from langchain_anthropic import ChatAnthropic
import time
from pydantic import BaseModel, Field
from langchain.output_parsers import PydanticOutputParser
from langchain.prompts import ChatPromptTemplate, SystemMessagePromptTemplate, HumanMessagePromptTemplate
import json
import openpyxl
from email.mime.text import MIMEText
import smtplib
import logging
from email.utils import formataddr

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

os.getenv("GROQ_API_KEY")
os.getenv("TAVILY_API_KEY")
os.getenv("ANTHROPIC_API_KEY")

def initialize_llm():
    return ChatGroq(model="llama-3.1-70b-versatile")

search = TavilySearchResults()

def read_excel_data(file_path, start_row):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        if row[2] != "email not found" or "Not provided":  # Assuming email is in the 3rd column (index 2)
            data.append({
                "meta_description": row[3],
                "username": row[0],
                "page_name": row[1],
                "email": row[2]
            })
    return data

def generate_email_body(page_name: str, meta_description: str):
    try:
        email_body_system_prompt = """
        Write a personalized email body offering AI automation services based on the provided data. 

        Follow these guidelines:
        
        **Strictly Dont Use Any Emojis and Dont Write any subject line or heading and dont write [Name] anywhere you have to give first name of the person you are writing to.**
        1. Limit to 3 short paragraphs, each 2-3 lines long not even a line more than that and not even a paragraph more than that try to cover up everything eing said in these 3 paragraphs .
        2. Paragraph 1: Start with a specific example of how AI voice agent can help their business that must include (voice agents capturing lead information).
        3. Paragraph 2: List 2-3 specific tasks AI agents can handle in their industry. Include a concrete benefit with numbers.
        5. Final paragraph: End with a clear call-to-action offering a call or custom demo and paste my website link in a new line https://optimoforge-ai.odoo.com/ .
        6.And at the end say Best Regards Abdullah , OptimoForge AI.
        6. Keep the tone professional yet conversational .
        7. Ensure each paragraph serves a distinct purpose within the limited space.


        Sample Example :
        Hi [Name],

        I hope this message finds you well. Your profile caught my eye, and I was impressed by the stunning properties you've listed and and I couldn't help but reach out.

        I'm Abdullah from Optimo Forge AI, and we're on a mission to revolutionize real estate with AI. Imagine having a voice agent that handles property inquiries and captures crucial leads. It can gather names, numbers, emails, pricing preferences, and location interests - all ready for personalized follow-ups It can also make inspection bookings or schedule inspections, freeing up your team to focus on what they do best - building relationships and closing deals.

        Our automation tools also streamline property listing updates across multiple platforms. This ensures your inventory is always current and can reduce manual data entry by up to 40%. As an experienced realtor, you probably know better than anyone how time-consuming these tasks can be.

        But here's where it gets really interesting: Our AI agents can handle tasks like scheduling viewings, answering FAQs, and even initial property valuations. This frees up your team to focus on what they do best - building relationships and closing deals. We've seen agencies boost efficiency by 30% and increase lead conversion rates by 25%.

        I'd love to show you how this could work for your agency. Are you free for a quick chat next week? Alternatively, we could send you a custom demo tailored to your needs.

        I'm excited about the possibility of working together you can reach out to us on our website https://optimoforge-ai.odoo.com/ . Let me know if you'd like to explore this further!

        Best regards, Abdllah
        OptimoForge AI 
        """

        email_body_prompt = """
        Client Name: {page_name}
        Business intro: {meta_description}

        Craft a personalized cold email pitch for EroHades AI services based on the given information and following the guidelines.
        """

        email_body_prompt_template = ChatPromptTemplate.from_messages([
            SystemMessagePromptTemplate.from_template(email_body_system_prompt),
            HumanMessagePromptTemplate.from_template(email_body_prompt)
        ])

        formatted_email_body_prompt = email_body_prompt_template.format_messages(
            page_name=page_name,
            meta_description=meta_description
        )

        groq_llm = initialize_llm()
        email_body_response = groq_llm.invoke(formatted_email_body_prompt)

        return email_body_response.content
    except Exception as e:
        logger.error(f"Error in generating email body: {str(e)}")
        return None

def generate_subject_line(email_body: str, business_info: str):
    try:
        subject_line_system_prompt = """
        You are an expert in selecting attention-grabbing email subject lines. Your task is to choose the most appropriate subject line from a given list and personalize it for the target client or company.

        Guidelines:
        - Do not use emojis
        - Provide only the selected subject line, without any explanations
        - Replace the Client's Name/Company Name in the chosen line with the  Client's First name or Company name

        Choose one line from the following options and personalize it by replacing Client's Name/Company Name with Clients First Name or Company Name mentioned in the email body that will be provided  :

        - Client's Name/Company Name AI assistant or human clone?
        - Client's Name/Company Name Slash paperwork, boost property sales with AI
        - Client's Name/Company Name Less busywork, more handshakes
        - AI is transforming real estate. Are you in, Client's Name/Company Name ?
        Output: The selected subject line with the client's name or company name replacing the original company name.
        """

        subject_line_prompt = """
        Email Body:
        {email_body}

        Business Info:
        {business_info}

        Create a catchy subject line for the email based on the given email body and business info.
        """

        subject_line_prompt_template = ChatPromptTemplate.from_messages([
            SystemMessagePromptTemplate.from_template(subject_line_system_prompt),
            HumanMessagePromptTemplate.from_template(subject_line_prompt)
        ])

        formatted_subject_line_prompt = subject_line_prompt_template.format_messages(
            email_body=email_body,
            business_info=business_info
        )

        groq_llm = initialize_llm()
        subject_line_response = groq_llm.invoke(formatted_subject_line_prompt)

        return subject_line_response.content
    except Exception as e:
        logger.error(f"Error in generating subject line: {str(e)}")
        return None
class CheckOutput(BaseModel):
    """Used to check if the output is relevant to our services"""
    is_relevant: str = Field(description="Check if the output is relevant to our AI automation services and follows the given guidelines. Give a one word answer either 'Yes' OR 'No' nothing else is accepted")

output_checker_parser = PydanticOutputParser(pydantic_object=CheckOutput)

def check_output_relevance(email_body: str, subject_line: str):
    try:
        checker_system_prompt = """
        **Guidelines for email body:
        1. Should be 3 short paragraphs, each 2-3 lines long.
        2. Must include specific examples of how AI voice agents can help the business.
        3. Should list 2-3 specific tasks AI agents can handle in their industry.
        4. Must include concrete benefits with numbers.
        5. Should end with a clear call-to-action offering a call or custom demo.
        6. Tone should be professional yet conversational.
        7. No emojis should be used.
        8. This website link must be attached in email https://optimoforge-ai.odoo.com/ .

        **Guidelines for subject line:
        1. Should be attention-grabbing and relevant to AI automation services.
        2. Must not use emojis.
        3. Should be personalized with the client's First Name or company name.

        Only give a binary score of either 'Yes' if relevant and following guidelines, or 'No' if not.
        """

        checker_human_prompt = """
        Subject Line: {subject_line}
        
        Email Body:
        {email_body}

        Are both the subject line and email body relevant to our AI automation services and do they follow the guidelines?

        {format_instructions}
        """

        checker_prompt_template = ChatPromptTemplate.from_messages([
            SystemMessagePromptTemplate.from_template(checker_system_prompt),
            HumanMessagePromptTemplate.from_template(checker_human_prompt)
        ])

        formatted_checker_prompt = checker_prompt_template.format_messages(
            subject_line=subject_line,
            email_body=email_body,
            format_instructions=output_checker_parser.get_format_instructions()
        )

        groq_llm = initialize_llm()
        checker_response = groq_llm.invoke(formatted_checker_prompt)

        if isinstance(checker_response, dict):
            is_relevant = checker_response.get('is_relevant', 'Unknown')
        elif isinstance(checker_response, CheckOutput):
            is_relevant = checker_response.is_relevant
        elif hasattr(checker_response, 'content'):
            # Handle AIMessage object
            content = checker_response.content.lower()
            is_relevant = 'Yes' if 'yes' in content else 'No'
        else:
            logger.error(f"Unexpected result type: {type(checker_response)}")
            return 'No'

        logger.info(f"Content Relevance: {is_relevant}")
        return is_relevant

    except Exception as e:
        logger.error(f"Error in checking output relevance: {str(e)}")
        return 'No'

def send_email(smtp_server, smtp_port, email_sender, app_password, email_recipient, subject_line, email_body):
    try:
        # Prepare the email
        message = MIMEText(email_body)
        message['From'] = formataddr(("OptimoForge AI", email_sender))
        message['to'] = email_recipient
        message['subject'] = subject_line

        # Set up the SMTP server
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()  # Start TLS encryption
            server.login(email_sender, app_password)  # Log in with your email and app password
            server.sendmail(email_sender, email_recipient, message.as_string())

        logger.info(f"Email sent to {email_recipient}")
    except Exception as e:
        logger.error(f"An error occurred while sending email to {email_recipient}: {e}")
        return None

def log_irrelevant_data(data, output_file='irrelevant_data.xlsx'):
    try:
        # Check if the file exists
        if os.path.exists(output_file):
            # If it exists, load the existing workbook
            wb = openpyxl.load_workbook(output_file)
            ws = wb.active
        else:
            # If it doesn't exist, create a new workbook and add headers
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["username", "page_name", "email", "meta_description", "generated_email", "generated_subject"])

        # Append the new row
        ws.append([
            data['username'],
            data['page_name'],
            data['email'],
            data['meta_description'],
            data['generated_email'],
            data['generated_subject']
        ])

        # Save the workbook
        wb.save(output_file)
        logger.info(f"Irrelevant data appended to {output_file}")
    except Exception as e:
        logger.error(f"Error logging irrelevant data: {str(e)}")

def main():
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    email_sender = "optimoforgeai@gmail.com"
    app_password = "xyrm ydbm zcyy usms"  

    excel_file_path = 'C:/Users/abdul/OneDrive/Documents/Leads.xlsx'
    start_row = int(input("Enter the starting row number: "))
    data = read_excel_data(excel_file_path, start_row)

    for row in data:
        email_body = generate_email_body(row['page_name'], row['meta_description'])
        if email_body:
            subject_line = generate_subject_line(email_body, f"{row['page_name']} - {row['meta_description']}")
            
            if subject_line:
                relevance_check = check_output_relevance(email_body, subject_line)
                if relevance_check.lower() == 'yes':
                    try:
                        send_email(smtp_server, smtp_port, email_sender, app_password, row['email'], subject_line, email_body)
                        logger.info(f"Email sent successfully to {row['email']}")
                    except Exception as e:
                        logger.error(f"Error sending email to {row['email']}: {e}")
                else:
                    logger.warning(f"Generated content for {row['email']} was deemed irrelevant. Logging for review.")
                    row['generated_email'] = email_body
                    row['generated_subject'] = subject_line
                    log_irrelevant_data(row)
            else:
                logger.error(f"Failed to generate subject line for {row['email']}")
        else:
            logger.error(f"Failed to generate email body for {row['email']}")

    logger.info("Email campaign completed!")

if __name__ == "__main__":
    main()   